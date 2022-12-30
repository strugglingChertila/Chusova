import csv
import datetime
from datetime import datetime
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

currency_to_rub = {'AZN': 35.68, 'BYR': 23.91, 'EUR': 59.90, 'GEL': 21.74, 'KGS': 0.76, 'KZT': 0.13, 'RUR': 1,
                   'UAH': 1.64, 'USD': 60.66, 'UZS': 0.0055}


class UsersInput:
    def __init__(self):
        self.compiled_file = input('Введите название файла: ')
        self.position_title = input('Введите название профессии: ')
        self.compiled_file = self.validate_file_name(self.compiled_file)
        self.position_title = self.check_profession_name(self.position_title)

    @staticmethod
    def validate_file_name(compiled_file):
        if compiled_file == '' or '.' not in compiled_file:
            print('Некорректное название файла')
            exit()
        return compiled_file

    @staticmethod
    def check_profession_name(position_title):
        if position_title == '':
            print('Некорректное название профессии')
            exit()
        return position_title


class DataSet:
    def __init__(self, compiled_file):
        self.reader = []
        for row in csv.reader(open(compiled_file, encoding='utf_8_sig')):
            self.reader += [row]
        if len(self.reader) == 0:
            print('Пустой файл')
            exit()
        self.columns_names = self.reader[0]
        self.vacancies_data = []
        for row in self.reader[1:]:
            if len(row) == len(self.columns_names) and row.count('') == 0:
                self.vacancies_data += [row]
        if len(self.vacancies_data) == 0:
            print('Нет данных')
            exit()


class Vacancy:
    name: str
    salary_from: int or float
    salary_to: int or float
    salary_currency: str
    area_name: str
    published_at: str
    salary: str

    def __init__(self, position):

        for key, value in position.items():
            self.__setattr__(key, self.formatter(key, value))

    @staticmethod
    def formatter(key, value):
        if key in ['salary_from', 'salary_to']:
            return float(value)
        if key == 'published_at':
            return int(datetime.strptime(value, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
        return value


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency


class SalaryDict:
    def __init__(self):
        self.salary_dict = {}
        self.__average_salary_dict = {}

    def append_salary(self, key, salary):
        if self.salary_dict.get(key) is None:
            self.salary_dict[key] = []
        return self.salary_dict[key].append(salary)

    def calculate_average_salary(self):
        for key, value in self.salary_dict.items():
            self.__average_salary_dict[key] = int(mean(value))
        return self.__average_salary_dict


class AmountChecker:
    def __init__(self):
        self.length = 0
        self.amount_dict = {}
        self.big_towns_arr = []
        self.prevailing_dict = {}

    def update_amount(self, key):
        if self.amount_dict.get(key) is None:
            self.amount_dict[key] = 0
        self.amount_dict[key] += 1
        self.length += 1
        return

    def calculate_proportion(self):
        proportion_dict = {}
        for key, value in self.amount_dict.items():
            found_proportion = value / self.length
            if found_proportion >= 0.1:
                self.big_towns_arr.append(key)
                proportion_dict[key] = round(found_proportion, 4)
        reversed_dict = dict(sorted(proportion_dict.items(), key=lambda row: row[1], reverse=True))
        self.prevailing_dict = {x: reversed_dict[x] for x in list(reversed_dict)[:10]}
        return


class Constractor:
    def __init__(self):
        self.year_salary = SalaryDict()
        self.year_vacancy_amount = AmountChecker()
        self.year_vacancy_salary = SalaryDict()
        self.year_position_vacancy_amount = AmountChecker()
        self.town_salary = SalaryDict()
        self.town_job_rate = AmountChecker()

    def compile_data(self, vacancies, place):
        self.calculate_stat_values(place, vacancies)
        if self.year_vacancy_salary.salary_dict == {}:
            self.year_vacancy_salary.salary_dict = {x: [0] for x in self.year_salary.salary_dict.keys()}
        elif self.year_vacancy_salary.salary_dict != {} and len(
                list(self.year_salary.calculate_average_salary().keys())) != len(
            list(self.year_vacancy_salary.calculate_average_salary().keys())):
            self.set_key_to_zero()
        if self.year_position_vacancy_amount.amount_dict == {}:
            self.year_position_vacancy_amount.amount_dict = {x: 0 for x in self.year_vacancy_amount.amount_dict.keys()}
        elif self.year_position_vacancy_amount.amount_dict != {} and len(
                list(self.year_vacancy_amount.amount_dict.keys())) != len(
            list(self.year_position_vacancy_amount.amount_dict.keys())):
            self.set_position_key_to_zero()
        self.town_salary, del_for_towns = self.calculate_highest_average_salary(self.town_salary)
        self.town_job_rate.calculate_proportion()
        self.town_job_rate = self.find_highest_town_rating(self.town_job_rate)
        self.town_job_rate = dict((x, y) for x, y in self.town_job_rate)
        return self.year_salary.calculate_average_salary(), self.year_vacancy_amount.amount_dict, \
               self.year_vacancy_salary.calculate_average_salary(), self.year_position_vacancy_amount.amount_dict, \
               self.town_salary, self.town_job_rate

    def set_position_key_to_zero(self):
        for key in list(self.year_vacancy_amount.amount_dict.keys()):
            if key not in list(self.year_position_vacancy_amount.amount_dict.keys()):
                self.year_position_vacancy_amount.amount_dict[key] = 0

    def set_key_to_zero(self):
        for key in list(self.year_salary.calculate_average_salary().keys()):
            if key not in list(self.year_vacancy_salary.calculate_average_salary().keys()):
                self.year_vacancy_salary.calculate_average_salary()[key] = 0

    def calculate_stat_values(self, place, vacancies):
        for position in vacancies:
            salary_given = (position.salary_from + position.salary_to) / 2 * currency_to_rub[position.salary_currency]
            self.year_salary.append_salary(position.published_at, salary_given)
            self.year_vacancy_amount.update_amount(position.published_at)
            self.town_salary.append_salary(position.area_name, salary_given)
            self.town_job_rate.update_amount(position.area_name)
            if place in position.name:
                self.year_vacancy_salary.append_salary(position.published_at, salary_given)
                self.year_position_vacancy_amount.update_amount(position.published_at)

    @staticmethod
    def calculate_highest_average_salary(list_all_salary):
        average_salary_values = []
        town_tracker = {}
        for i in range(len(list_all_salary.salary_dict)):
            town = list(list_all_salary.salary_dict)[i]
            town_tracker[town] = len(list(list_all_salary.salary_dict.values())[i])
            aver = int(sum(list(list_all_salary.salary_dict.values())[i]) /
                       len(list(list_all_salary.salary_dict.values())[i]))
            average_salary_values.append((town, aver))

        del_for_towns = []
        del_town_index = []
        for i in range(len(town_tracker.items())):
            town = list(town_tracker)[i]
            percentage = round(100 * int(list(town_tracker.values())[i]) / sum(town_tracker.values()), 1)
            if percentage < 1 or town == 'Россия':
                del_for_towns.append((town, list(town_tracker.values())[i]))
                del_town_index.append(i)

        for i in reversed(range(len(del_for_towns))):
            del town_tracker[del_for_towns[i][0]]
            del average_salary_values[del_town_index[i]]

        highest_average_salary = dict(sorted(average_salary_values, key=lambda row: row[1], reverse=True))
        significant_salaries = {}
        for key, value in highest_average_salary.items():
            significant_salaries[key] = value
        return {x: significant_salaries[x] for x in list(significant_salaries)[:10]}, del_for_towns

    @staticmethod
    def find_highest_town_rating(town_job_rate):
        del_for_towns = []
        for i in reversed(range(len(del_for_towns))):
            del town_job_rate.amount_dict[del_for_towns[i][0]]

        proportion_dict = {}

        for key, value in town_job_rate.amount_dict.items():
            found_proportion = value / town_job_rate.length
            if found_proportion >= 0.01:
                proportion_dict[key] = round(found_proportion, 4)

        reversed_dict = sorted(proportion_dict.items(), key=lambda row: row[1], reverse=True)
        return reversed_dict[:10]


class CreateReport:
    def __init__(self):
        self.wb = Workbook()
        self.sheet1 = self.wb.active
        self.sheet1.title = 'Статистика по годам'
        self.sheet2 = self.wb.create_sheet('Статистика по городам')

    def create_excel_sheet(self, data, place):
        year_salary = data[0]
        year_vacancy_amount = data[1]
        year_vacancy_salary = data[2]
        year_position_vacancy_amount = data[3]
        town_salary = data[4]
        town_job_rate = data[5]

        titles_1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {place}',
                    'Количество вакансий', f'Количество вакансий - {place}']
        titles_2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']

        for i, name in enumerate(titles_1):
            self.sheet1.cell(row=1, column=(i + 1), value=name).font = Font(bold=True)
        for year, value in year_salary.items():
            self.sheet1.append(
                [year, value, year_vacancy_salary[year], year_vacancy_amount[year], year_position_vacancy_amount[year]])

        for i, name in enumerate(titles_2):
            self.sheet2.cell(row=1, column=(i + 1), value=name).font = Font(bold=True)
        for i in range(len(list(town_salary.keys()))):
            self.sheet2.append(
                [list(town_salary.keys())[i], list(town_salary.values())[i], list(town_job_rate.keys())[i],
                 list(town_job_rate.values())[i]])

        cell_border = Side(border_style='thin', color='000000')
        self.place_border(self.sheet1, cell_border)
        self.place_border(self.sheet2, cell_border)
        self.sheet2.insert_cols(3)
        self.sheet2.column_dimensions['C'].width = 2

        self.calculate_column_width(self.sheet1)
        self.calculate_column_width(self.sheet2)

        for i in range(2, len(self.sheet2['E']) + 1):
            self.sheet2[f'E{i}'].number_format = FORMAT_PERCENTAGE_00

        self.wb.save('report.xlsx')

    @staticmethod
    def place_border(ws, cell_border):
        for cell in ws._cells.values():
            cell.border = Border(top=cell_border, bottom=cell_border, left=cell_border, right=cell_border)

    @staticmethod
    def calculate_column_width(ws):
        dimension_dict = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dimension_dict[cell.column_letter] = max(
                        (dimension_dict.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dimension_dict.items():
            ws.column_dimensions[col].width = value + 2


def generate_output(data_vacancies, position_title):
    compiled_vacancies_arr = []
    for compilation in data_vacancies:
        compilation = Vacancy(dict(zip(column_headers, compilation)))
        compiled_vacancies_arr.append(compilation)
    data = Constractor()
    data = data.compile_data(compiled_vacancies_arr, position_title)

    print(f'Динамика уровня зарплат по годам: {data[0]}')
    print(f'Динамика количества вакансий по годам: {data[1]}')
    print(f'Динамика уровня зарплат по годам для выбранной профессии: {data[2]}')
    print(f'Динамика количества вакансий по годам для выбранной профессии: {data[3]}')
    print(f'Уровень зарплат по городам (в порядке убывания): {data[4]}')
    print(f'Доля вакансий по городам (в порядке убывания): {data[5]}')

    return data


users_input = UsersInput()
requested_data = DataSet(users_input.compiled_file)
column_headers, vacancies_data = requested_data.columns_names, requested_data.vacancies_data
output_data = generate_output(vacancies_data, users_input.position_title)
generated_report = CreateReport()
generated_report.create_excel_sheet(output_data, users_input.position_title)
